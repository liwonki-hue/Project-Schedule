import streamlit as st
import pandas as pd
import numpy as np
import datetime
import io
import plotly.express as px
import plotly.graph_objects as go
import streamlit.components.v1 as components


def create_revision_excel(rev_df, mc_target):
    """Generate A3 landscape Excel with Milestone table + Gantt chart.
    Uses current rev_df so it auto-updates when sidebar dates change.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Revision Milestone Forecast"

    # A3 Landscape
    ws.page_setup.paperSize = 8
    ws.page_setup.orientation = "landscape"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # Styles
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    center   = Alignment(horizontal="center", vertical="center")
    alt_fill = PatternFill("solid", fgColor="EFF6FF")
    red_fill = PatternFill("solid", fgColor="FECACA")
    thin = Side(style="thin", color="D1D5DB")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title (row 1)
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"Revision Milestone Forecast — GT #11 + GT #12    MC Target: {mc_target}"
    t.font  = Font(bold=True, size=13)
    t.alignment = center
    ws.row_dimensions[1].height = 24

    # Table headers (row 3)
    HEADERS = ["#", "Area", "Work Teams", "Handover",
               "Piping (DI)", "Support (EA)", "Pressure Test Finish", "Float (D)"]
    COL_W   = [4, 24, 12, 14, 12, 12, 22, 10]
    for ci, (h, w) in enumerate(zip(HEADERS, COL_W), 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = center; c.border = bdr
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 18

    def to_date(v):
        if isinstance(v, datetime.datetime): return v.date()
        if isinstance(v, datetime.date):     return v
        return datetime.datetime.strptime(str(v), "%Y-%m-%d").date()

    ref_date = min(to_date(r["Handover"]) for _, r in rev_df.iterrows())
    n = len(rev_df)

    # Data rows (rows 4 … 3+n)
    for i, (_, row) in enumerate(rev_df.iterrows(), 1):
        dr = 3 + i
        hd = to_date(row["Handover"])
        ft = to_date(row["Pressure Test Finish"])
        fv = int(row["Float"])
        vals = [i, row["Area"], round(float(row["Work Teams"]), 1),
                hd, int(row["Piping(DI)"]), int(row["Support(EA)"]), ft, fv]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=dr, column=ci, value=v)
            c.alignment = center; c.border = bdr
            if i % 2 == 0: c.fill = alt_fill
            if ci in (4, 7): c.number_format = "YYYY-MM-DD"
            if ci == 8 and fv < 0:
                c.fill = red_fill
                c.font = Font(color="991B1B", bold=True)
        ws.row_dimensions[dr].height = 16

    # Total row
    tr = 3 + n + 1
    for ci in range(1, 9):
        ws.cell(row=tr, column=ci).border = bdr
    ws.cell(row=tr, column=2, value="TOTAL").font    = Font(bold=True)
    ws.cell(row=tr, column=3, value=round(float(rev_df["Work Teams"].sum()), 1)).font = Font(bold=True)
    ws.cell(row=tr, column=5, value=int(rev_df["Piping(DI)"].sum())).font             = Font(bold=True)
    ws.cell(row=tr, column=6, value=int(rev_df["Support(EA)"].sum())).font            = Font(bold=True)
    ws.cell(row=tr, column=8, value=int(rev_df["Float"].min())).font                  = Font(bold=True)

    # Gantt source data (columns J-L, same sheet)
    GC = 10
    ws.cell(row=3, column=GC,   value="Area")
    ws.cell(row=3, column=GC+1, value="Offset")
    ws.cell(row=3, column=GC+2, value="Duration")
    ws.column_dimensions["J"].width = 22
    ws.column_dimensions["K"].width = 10
    ws.column_dimensions["L"].width = 10
    for i, (_, row) in enumerate(rev_df.iterrows(), 1):
        dr = 3 + i
        hd = to_date(row["Handover"])
        ft = to_date(row["Pressure Test Finish"])
        ws.cell(row=dr, column=GC,   value=row["Area"])
        ws.cell(row=dr, column=GC+1, value=(hd - ref_date).days)
        ws.cell(row=dr, column=GC+2, value=(ft - hd).days)

    # Gantt bar chart
    chart = BarChart()
    chart.type      = "bar"
    chart.grouping  = "stacked"
    chart.overlap   = 100
    chart.title     = f"Revision Construction Path  (MC: {mc_target})"
    chart.style     = 2
    chart.height    = 15
    chart.width     = 25

    cats = Reference(ws, min_col=GC,   min_row=4, max_row=3+n)
    off  = Reference(ws, min_col=GC+1, min_row=3, max_row=3+n)
    dur  = Reference(ws, min_col=GC+2, min_row=3, max_row=3+n)
    chart.add_data(off, titles_from_data=True)
    chart.add_data(dur, titles_from_data=True)
    chart.set_categories(cats)

    # Offset series → white (invisible), Duration series → green
    try:
        from openpyxl.drawing.fill import ColorChoice
        chart.series[0].graphicalProperties.solidFill = "FFFFFF"
        chart.series[1].graphicalProperties.solidFill = "22C55E"
    except Exception:
        pass  # colour setting is cosmetic; skip if API differs

    try:
        from openpyxl.chart.axis import Scaling
        chart.y_axis.scaling = Scaling(orientation="maxMin")
    except Exception:
        pass  # Area order fallback

    chart.x_axis.title = f"Days from {ref_date.strftime('%Y-%m-%d')}"
    chart.y_axis.title = ""

    ws.add_chart(chart, "I1")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()   # return bytes, not BytesIO

# --- Global Layout Config ---
st.set_page_config(page_title="Strategic Piping Management", layout="wide", page_icon="🏗️")

# --- UI Aesthetics ---
st.markdown("""
    <style>
    /* Global Compactness */
    [data-testid="stVerticalBlock"] > div { padding-top: 0px !important; padding-bottom: 0px !important; gap: 0.1rem !important; }
    .stApp { background-color: white !important; }
    
    /* Remove Backgrounds & Borders */
    .report-card, .flow-box, .solution-card { 
        background: transparent !important; 
        border: none !important; 
        padding: 5px !important;
        color: black !important;
    }
    
    /* Report Font Adjustments: Titles 1.1x larger than baseline */
    .report-card h2, .flow-box h2 { font-size: 1.45rem !important; margin-top: 5px !important; margin-bottom: 5px !important; color: #1e293b !important; }
    .report-card h3, .flow-box h3 { font-size: 1.25rem !important; margin-top: 5px !important; margin-bottom: 2px !important; color: #475569 !important; }
    
    .report-card p, .report-card li, .flow-box p, .flow-box li, .solution-card p, .solution-card li { 
        font-size: 1.1rem !important; 
        line-height: 1.1 !important; 
        margin-bottom: 5pt !important; 
    }
    .report-card ul, .report-card ol, .flow-box ul, .solution-card ul { 
        margin-top: 5pt !important; 
        margin-bottom: 5pt !important; 
    }
    
    /* Enhanced Table Styling */
    .report-card table { width: 50% !important; border-collapse: collapse !important; margin-top: 15px !important; }
    .report-card th { background-color: #f8fafc !important; border-bottom: 2px solid #cbd5e1 !important; padding: 12px !important; }
    .report-card td { padding: 12px !important; border-bottom: 1px solid #e2e8f0 !important; }
    
    /* Lower Section Tables: Shrink only the Essential Systems table to 80% */
    .compact-section table { 
        width: 80% !important; 
        margin-left: 0 !important; 
        margin-right: auto !important;
    }
    
    /* Center align table headers and cells via CSS hack */
    .stDataFrame [data-testid="stTable"] th { text-align: center !important; vertical-align: middle !important; font-size: 1.15rem !important; }
    .stDataFrame [data-testid="stTable"] td { text-align: center !important; vertical-align: middle !important; font-size: 1.1rem !important; }
    .stDataFrame [data-testid="stMetric"] { text-align: center; } /* Also center metrics */
    
    /* Metric compression - Scale up by 1.3x */
    [data-testid="stMetricValue"] { font-size: 1.6rem !important; font-weight: 700 !important; }
    [data-testid="stMetricLabel"] { font-size: 1.2rem !important; }
    /* Essential Systems Table Optimization */
    [data-testid="stTable"] td {
        font-size: 0.95rem !important;
        vertical-align: middle !important;
    }
    /* Index, System, Area, Description, Criticality, Related Equipment: Keep in one line */
    [data-testid="stTable"] td:nth-child(1), 
    [data-testid="stTable"] td:nth-child(2),
    [data-testid="stTable"] td:nth-child(3),
    [data-testid="stTable"] td:nth-child(4),
    [data-testid="stTable"] td:nth-child(5),
    [data-testid="stTable"] td:nth-child(6) {
        white-space: nowrap !important;
    }
    /* Allow wrapping only for Remark(7th) to absorb extra space */
    [data-testid="stTable"] td:nth-child(7) {
        white-space: normal !important;
        min-width: 120px;
    }

    [data-testid="stMetric"] { margin-bottom: -1rem !important; text-align: center; }
    
    .solution-card {
        background: #eef2ff !important; /* Indigo-tinted background */
        border-left: 6px solid #4338ca !important; /* Indigo border */
        padding: 20px !important;
        border-radius: 8px !important;
        color: #1e1b4b !important; /* Deep indigo text */
        line-height: 1.6 !important;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }
    .solution-title {
        font-size: 1.6rem !important;
        font-weight: 800 !important;
        margin-bottom: 12px !important;
        display: block;
        color: #312e81 !important;
    }
    .solution-content {
        font-size: 1.25rem !important;
        font-weight: 700 !important;
    }

    [data-testid="stHeader"] { display: none; }
    .block-container { padding-top: 0.5rem !important; padding-bottom: 0.5rem !important; padding-left: 2rem !important; padding-right: 2rem !important; }
    
    /* Sidebar Spacing */
    [data-testid="stSidebar"] { background-color: #f8fafc !important; }
    [data-testid="stSidebar"] [data-testid="stVerticalBlock"] { padding-top: 25px !important; gap: 0.6rem !important; }
    [data-testid="stSidebar"] .stMarkdown { margin-bottom: 5px !important; }
    [data-testid="stSidebar"] .stDateInput, [data-testid="stSidebar"] .stSlider, [data-testid="stSidebar"] .stNumberInput { margin-bottom: 8px !important; }
    [data-testid="stSidebar"] [data-testid="stWidgetLabel"] p { font-weight: 700 !important; font-size: 0.95rem !important; color: #334155 !important; }
</style>
""", unsafe_allow_html=True)

# --- Master Data ---
AREA_DATA = {
    "GT #11 Installation": {"EA": 895, "DI": 4115, "Lag": 0, "Status": "Alignment (50%)"}, # Added AS/HW (+1649 DI, +518 EA)
    "HRSG #11 PR": {"EA": 173, "DI": 1131, "Lag": 2, "Status": "In Progress"},
    "Main Building Structure": {"EA": 1808, "DI": 10087, "Lag": 2, "Status": "Wait for MB Steel"}, # Added AS/HW (+1649 DI, +518 EA)
    "Pipe Rack #3": {"EA": 925, "DI": 4311, "Lag": 2, "Status": "Wait for Architectural"}, # Added AS/HW (+1649 DI, +518 EA)
    "Pipe Rack #4": {"EA": 1118, "DI": 5574, "Lag": 2, "Status": "Wait for Architectural"}, # Added AS/HW (+1649 DI, +518 EA)
    "Pipe Rack #5": {"EA": 266, "DI": 1740, "Lag": 2, "Status": "Wait for Architectural"},
    "Pipe Rack #6": {"EA": 173, "DI": 1131, "Lag": 2, "Status": "Wait for Architectural"},
    "Pipe Rack #7": {"EA": 221, "DI": 1445, "Lag": 2, "Status": "Wait for Architectural"}
}
TOTAL_DI = sum(d["DI"] for d in AREA_DATA.values())
TOTAL_EA = sum(d["EA"] for d in AREA_DATA.values())
MC_TARGET = datetime.date(2026, 9, 16)

# --- Revision: GT #12 Additional Areas (from Joint Master & Support Master) ---
REVISION_EXTRA_AREAS = {
    "HRSG #12 PR": {"EA": 353, "DI": 1617, "Lag": 2, "Status": "Not Started"},
    "GT #12 Installation": {"EA": 548, "DI": 2569, "Lag": 0, "Status": "Not Started"},
}
REVISION_AREA_DATA = {**AREA_DATA, **REVISION_EXTRA_AREAS}
REVISION_TOTAL_DI = sum(d["DI"] for d in REVISION_AREA_DATA.values())
REVISION_TOTAL_EA = sum(d["EA"] for d in REVISION_AREA_DATA.values())

# --- Display Order (Pipe Rack → MB → HRSG #11 → GT #11 → HRSG #12 → GT #12) ---
AREA_DISPLAY_ORDER = [
    "Pipe Rack #3", "Pipe Rack #4", "Pipe Rack #5", "Pipe Rack #6", "Pipe Rack #7",
    "Main Building Structure", "HRSG #11 PR", "GT #11 Installation",
    "HRSG #12 PR", "GT #12 Installation",
]
ORIGINAL_DISPLAY_ORDER = AREA_DISPLAY_ORDER[:8]  # GT #11 only (no GT #12 areas)

# --- Print Support ---
def add_print_button():
    components.html(
        """
        <script>
            function printReport() {
                window.parent.print();
            }
        </script>
        <div style="display: flex; justify-content: flex-end; align-items: center; height: 100%;">
            <button onclick="printReport()" style="
                background-color: #f8fafc;
                color: #475569;
                padding: 6px 14px;
                border: 1px solid #cbd5e1;
                border-radius: 6px;
                cursor: pointer;
                font-weight: 600;
                font-size: 13px;
                display: flex;
                align-items: center;
                gap: 6px;
                box-shadow: 0 1px 2px rgba(0,0,0,0.05);
            ">
                <span>🖨️</span> Print
            </button>
        </div>
        """,
        height=60
    )

# --- Sidebar Controls ---
st.sidebar.markdown('<h2 style="font-size: 1.35rem; margin-top: 0; margin-bottom: 5px;">📌 Simulation Panel</h2>', unsafe_allow_html=True)

st.sidebar.markdown('<hr style="margin: 10px 0;">', unsafe_allow_html=True)
st.sidebar.markdown('<p style="font-weight: 600; font-size: 1.05rem; margin-bottom: 8px;">Handover Dates</p>', unsafe_allow_html=True)
rel_dates = {}
for area in AREA_DATA.keys():
    d_rel = datetime.date(2026, 6, 1)
    if "#3" in area: d_rel = datetime.date(2026, 5, 20)
    if "HRSG #11 PR" in area: d_rel = datetime.date(2026, 7, 1)
    if "GT #11" in area: d_rel = datetime.date(2026, 7, 5)
    if "Main Building Structure" in area: d_rel = datetime.date(2026, 9, 1)
    rel_dates[area] = st.sidebar.date_input(area, d_rel)

st.sidebar.markdown('<hr style="margin: 5px 0;">', unsafe_allow_html=True)
st.sidebar.markdown('<p style="font-weight: 600; font-size: 1.05rem; margin-bottom: 8px; color: #4338ca;">Revision: GT #12 Areas</p>', unsafe_allow_html=True)
rev_extra_dates = {}
rev_extra_dates["HRSG #12 PR"] = st.sidebar.date_input("HRSG #12 PR", datetime.date(2026, 7, 1))
rev_extra_dates["GT #12 Installation"] = st.sidebar.date_input("GT #12 Installation", datetime.date(2026, 8, 1))

st.sidebar.markdown('<hr style="margin: 5px 0;">', unsafe_allow_html=True)
with st.sidebar.expander("Manpower Setup", expanded=True):
    total_manpower_teams = st.number_input("Piping Work Teams", 1, 100, 50)
    prod_di = st.slider("DI/Team-Day", 3.0, 30.0, 13.0)
    prod_ea_ratio = st.slider("EA/DI Ratio", 0.1, 1.0, 0.4)
    priority = st.slider("EP Priority (%)", 0, 100, 70)

# --- Simulation Logic ---
AREA_CAPS = {
    "Main Building Structure": 6, # Space constraint: Max 6 Work Teams
    "GT #11 Installation": 8,
    "HRSG #11 PR": 6,
    "Pipe Rack #3": 4,
    "Pipe Rack #4": 4,
    "Pipe Rack #5": 3,
    "Pipe Rack #6": 3,
    "Pipe Rack #7": 3,
}
REVISION_AREA_CAPS = {
    **AREA_CAPS,
    "HRSG #12 PR": 6,
    "GT #12 Installation": 8,
}

ep_work_teams = total_manpower_teams * (priority / 100)
prod_ea = prod_di * prod_ea_ratio
sim_results = []

# --- Resource Allocation Logic (Redistribution Algorithm) ---
def allocate_teams(total_teams, area_data, area_caps):
    assigned = {area: 0.0 for area in area_data}
    remaining_teams = total_teams
    
    # Iteratively distribute teams to areas not yet at capacity
    for _ in range(10): # Max 10 iterations to prevent infinite loops
        eligible_areas = [a for a in area_data if assigned[a] < area_caps.get(a, 4)]
        if not eligible_areas or remaining_teams < 0.1:
            break
            
        total_eligible_weight = sum(area_data[a]["DI"] for a in eligible_areas)
        if total_eligible_weight == 0: break
        
        starting_remaining = remaining_teams
        for area in eligible_areas:
            share = area_data[area]["DI"] / total_eligible_weight
            potential_assign = starting_remaining * share
            
            # How much more can this area take?
            space_left = area_caps.get(area, 4) - assigned[area]
            actual_assign = min(potential_assign, space_left)
            
            assigned[area] += actual_assign
            remaining_teams -= actual_assign
            
        if starting_remaining - remaining_teams < 0.05: break
    
    # Ensure minimum 1 team per active area if possible
    for area in assigned:
        if assigned[area] < 1.0 and remaining_teams > 0:
            assigned[area] = 1.0
            
    return assigned

assigned_teams_map = allocate_teams(ep_work_teams, AREA_DATA, AREA_CAPS)
prod_ea = prod_di * prod_ea_ratio
sim_results = []

# --- Calculate simulation with assigned teams ---
for area, data in AREA_DATA.items():
    piping_start = rel_dates[area] + datetime.timedelta(weeks=data["Lag"])
    assigned_teams = assigned_teams_map[area]
    
    # Calculate daily output for the area
    capacity_di = assigned_teams * prod_di
    piping_days = data["DI"] / capacity_di if capacity_di > 0 else 999
    
    piping_finish = piping_start + datetime.timedelta(days=int(piping_days))
    pt_finish = piping_finish + datetime.timedelta(weeks=2)
    flt = (MC_TARGET - pt_finish).days
    
    sim_results.append({
        "Area": area, 
        "Handover": rel_dates[area],
        "Schedule Status": data["Status"],
        "Work Teams": round(assigned_teams, 1),
        "Piping Finish": piping_finish,
        "Pressure Test Finish": pt_finish,
        "Piping(DI)": data["DI"], 
        "Support(EA)": data["EA"], 
        "Float": flt
    })

df = pd.DataFrame(sim_results)
max_finish = df["Pressure Test Finish"].max()
net_float = (MC_TARGET - max_finish).days
# Mathematical bottleneck
math_bottleneck = df.loc[df['Pressure Test Finish'].idxmax(), 'Area']
# User defined strategic bottleneck: Main Building Structure
strat_bottleneck = "Main Building Structure"

# --- Revision Simulation (Global scope: used by Dashboard + Revision tab) ---
rev_rel_dates = {**rel_dates, **rev_extra_dates}
rev_assigned_teams_map = allocate_teams(ep_work_teams, REVISION_AREA_DATA, REVISION_AREA_CAPS)
rev_sim_results = []
for area, data in REVISION_AREA_DATA.items():
    piping_start = rev_rel_dates[area] + datetime.timedelta(weeks=data["Lag"])
    assigned_teams = rev_assigned_teams_map[area]
    capacity_di = assigned_teams * prod_di
    piping_days = data["DI"] / capacity_di if capacity_di > 0 else 999
    piping_finish = piping_start + datetime.timedelta(days=int(piping_days))
    pt_finish = piping_finish + datetime.timedelta(weeks=2)
    flt = (MC_TARGET - pt_finish).days
    rev_sim_results.append({
        "Area": area,
        "Handover": rev_rel_dates[area],
        "Schedule Status": data["Status"],
        "Work Teams": round(assigned_teams, 1),
        "Piping Finish": piping_finish,
        "Pressure Test Finish": pt_finish,
        "Piping(DI)": data["DI"],
        "Support(EA)": data["EA"],
        "Float": flt
    })
rev_df = pd.DataFrame(rev_sim_results)
rev_max_finish = rev_df["Pressure Test Finish"].max()
rev_net_float = (MC_TARGET - rev_max_finish).days
rev_math_bottleneck = rev_df.loc[rev_df["Pressure Test Finish"].idxmax(), "Area"]

# --- Original Essential Systems (GT #11 only, pre-revision) ---
ORIGINAL_ESSENTIAL_SYSTEMS = [
    {"System": "Fuel Gas & FGH System", "Area": "MB / GT / FGH", "Description": "Supply Fuel Gas to GT", "DI": 2480, "EA": 610, "Criticality": "Mandatory", "Equipment": "FGSS System, Performance Heater & Final Filter", "Remark": ""},
    {"System": "Closed Cooling Water (CCW)", "Area": "MB / PR / PH#1 / FFC#1", "Description": "Supply & Return of CCW for equipment cooling down", "DI": 4120, "EA": 980, "Criticality": "Mandatory", "Equipment": "PH#1 & FFC#1 Integrated", "Remark": ""},
    {"System": "Instrument Air (IA)", "Area": "All Areas", "Description": "Supply Instrument Air to all plant", "DI": 1150, "EA": 290, "Criticality": "Operation", "Equipment": "Air Compressors, IA Dryers", "Remark": ""},
    {"System": "Aux. Steam & Hot Water (AS/HW)", "Area": "PR#3/4 / MB", "Description": "supply & return of Hot Water for GT Anti-Icing", "DI": 2650, "EA": 660, "Criticality": "Mandatory", "Equipment": "Aux. Boiler, Hot Water Supply Pump, Heat Exchanger", "Remark": ""},
    {"System": "Nitrogen System (N2)", "Area": "GT Area / PR", "Description": "Purge & Vent for Fuel Gas", "DI": 320, "EA": 90, "Criticality": "Mandatory", "Equipment": "N2 Storage Tank, N2 Supply System", "Remark": ""},
    {"System": "GT MISC (Vents)", "Area": "Main Building", "Description": "GT Lube Oil Mist & Fuel Gas Vent", "DI": 650, "EA": 180, "Criticality": "Mandatory", "Equipment": "GT Enclosure, Vent Fans", "Remark": ""},
    {"System": "Demineralized Water (DW)", "Area": "Water Treatment", "Description": "Supply Demi. Water to plant", "DI": 2000, "EA": 480, "Criticality": "Highest", "Equipment": "DW Tank & Pump Station", "Remark": "Under review for sourcing from other power plant (by INTEGRA)"},
]

# --- Revised Essential Systems (GT #11 + GT #12 + HRSG #12 PR quantities per system) ---
# GT#12 & HRSG#12 PR quantities distributed per system using EA-proportional DI estimation
# GT#12: 2,569 Field DI / 548 EA = 4.69 DI/EA | HRSG#12 PR: 1,617 Field DI / 353 EA = 4.58 DI/EA
ESSENTIAL_SYSTEMS = [
    {"System": "Fuel Gas & FGH System",
     "Area": "MB / GT / FGH / GT#12 / HRSG#12",
     "Description": "Supply Fuel Gas to GT #11 & GT #12",
     "DI": 2838, "EA": 687, "Criticality": "Mandatory",
     "Equipment": "FGSS System, Performance Heater & Final Filter",
     "Remark": ""},
    {"System": "Closed Cooling Water (CCW)",
     "Area": "MB / PR / PH#1 / FFC#1 / GT#12 / HRSG#12",
     "Description": "Supply & Return of CCW for equipment cooling (GT#11 & GT#12)",
     "DI": 4651, "EA": 1094, "Criticality": "Mandatory",
     "Equipment": "PH#1 & FFC#1 Integrated",
     "Remark": ""},
    {"System": "Instrument Air (IA)",
     "Area": "All Areas incl. GT#12 / HRSG#12",
     "Description": "Supply Instrument Air to all plant incl. GT#12 & HRSG#12",
     "DI": 1467, "EA": 358, "Criticality": "Operation",
     "Equipment": "Air Compressors, IA Dryers",
     "Remark": ""},
    {"System": "Aux. Steam & Hot Water (AS/HW)",
     "Area": "PR#3/4 / MB / GT#12 / HRSG#12",
     "Description": "Supply & Return of Hot Water for GT Anti-Icing (GT#11 & GT#12)",
     "DI": 3009, "EA": 737, "Criticality": "Mandatory",
     "Equipment": "Aux. Boiler, Hot Water Supply Pump, Heat Exchanger",
     "Remark": ""},
    {"System": "Nitrogen System (N2)",
     "Area": "GT Area / PR / GT#12 / HRSG#12",
     "Description": "Purge & Vent for Fuel Gas (GT#11 & GT#12)",
     "DI": 497, "EA": 128, "Criticality": "Mandatory",
     "Equipment": "N2 Storage Tank, N2 Supply System",
     "Remark": ""},
    {"System": "GT MISC (Vents)",
     "Area": "Main Building / GT#12",
     "Description": "GT Lube Oil Mist & Fuel Gas Vent (GT#11 & GT#12)",
     "DI": 1372, "EA": 334, "Criticality": "Mandatory",
     "Equipment": "GT Enclosure, Vent Fans",
     "Remark": ""},
    {"System": "Demineralized Water (DW)",
     "Area": "Water Treatment / GT#12 / HRSG#12",
     "Description": "Supply Demi. Water to plant incl. GT#12 & HRSG#12",
     "DI": 2301, "EA": 545, "Criticality": "Highest",
     "Equipment": "DW Tank & Pump Station",
     "Remark": "Under review for sourcing from other power plant (by INTEGRA)"},
]

# --- Presentation Layer ---
tab_dash, tab_rep1, tab_rev, tab_rep2 = st.tabs(["📊 Revision (GT #11)", "📋 Technical Report(1)", "📊 Revision (GT#11 + GT#12)", "📋 Technical Report(2)"])

with tab_dash:
    # Custom CSS for Table Alignment & Column Widths
    st.markdown("""
        <style>
            th, td { text-align: center !important; }
            /* Expand DI and EA columns (typically 5th and 6th in these tables) */
            th:nth-child(5), td:nth-child(5), 
            th:nth-child(6), td:nth-child(6) { 
                min-width: 95px !important; 
            }
        </style>
    """, unsafe_allow_html=True)
    
    t_col, p_col = st.columns([8.2, 1.8])
    with t_col:
        st.markdown('<h1 style="font-size: 1.8rem; margin: 0; padding: 0;">🏗️ Project Simulation Dashboard: GT #11 Early Power</h1>', unsafe_allow_html=True)
    with p_col:
        add_print_button()
    
    # Major Constraints Alert
    st.markdown(f"""
        <div style="background-color:#fff3e0; padding:15px; border-radius:10px; border-left:8px solid #ff9800; margin-bottom:10px;">
            <span style="font-size:1.15rem; font-weight:bold; color:#e65100;">
                🚨 Ultimate Bottleneck in timeline : {strat_bottleneck} (Structure Access for header piping installation) | Target Float: {net_float} Days
            </span>
        </div>
        <div style="background-color:#e0f2f1; padding:15px; border-radius:10px; border-left:8px solid #00897b; margin-bottom:20px;">
            <span style="font-size:1.15rem; font-weight:bold; color:#004d40;">
                ⚠️ Major Constraint: Power Receiving must be completed before CCW/HW Pump operation. 
                Energizing requires FF (Fire Fighting) & HVAC installation to prevent overheating/fire.
            </span>
        </div>
    """, unsafe_allow_html=True)
    
    # KPI Grid
    m1, m2, m3, m4, m5, m6 = st.columns(6)
    m1.metric("Piping Vol. (Field)", f"{TOTAL_DI:,} DI")
    m2.metric("Support Vol.", f"{TOTAL_EA:,} EA")
    m3.metric("Avg Field DI/DAY", f"{prod_di} DI")
    m4.metric("Avg Support EA/DAY", f"{prod_ea:.1f} EA")
    m5.metric("Target Float", f"{net_float} D", delta_color="normal" if net_float >= 0 else "inverse")
    m6.metric("Total Teams", f"{ep_work_teams:.1f}")
    
    st.divider()
    
    left, right = st.columns([4.5, 5.5])
    with left:
        st.subheader("📊 Detailed Milestone Forecast")
        st.markdown('<div style="height: 40px;"></div>', unsafe_allow_html=True)
        # Reorder by display order (GT #11 only — 8 areas)
        df_ordered = df.set_index("Area").reindex(ORIGINAL_DISPLAY_ORDER).reset_index()
        df_view = df_ordered.copy()
        df_view["Work Teams Temp"] = df_view["Work Teams"].map(lambda x: f"{x:.1f}")
        for c in ["Handover", "Pressure Test Finish"]:
            df_view[c] = df_view[c].apply(lambda x: x.strftime('%Y-%m-%d'))

        display_cols = ["Area", "Work Teams Temp", "Handover", "Piping(DI)", "Support(EA)", "Pressure Test Finish", "Float"]
        df_display = df_view[display_cols].rename(columns={"Work Teams Temp": "Work Teams"})

        # --- Add Total Row ---
        total_row = pd.DataFrame([{
            "Area": "TOTAL",
            "Work Teams": f"{df['Work Teams'].sum():.1f}",
            "Handover": "",
            "Piping(DI)": int(df["Piping(DI)"].sum()),
            "Support(EA)": int(df["Support(EA)"].sum()),
            "Pressure Test Finish": "",
            "Float": int(df["Float"].min())
        }])
        df_display = pd.concat([df_display, total_row], ignore_index=True)
        new_index = [str(i+1) for i in range(len(df_display)-1)] + [""]
        df_display.index = new_index

        st.table(df_display)

    with right:
        st.subheader("📅 Construction Path")
        # Use display-ordered data for chart (GT #11 only)
        chart_df = df.set_index("Area").reindex(ORIGINAL_DISPLAY_ORDER).reset_index()
        fig = px.timeline(chart_df, x_start="Handover", x_end="Pressure Test Finish", y="Area",
                          color="Float", color_continuous_scale=["#ef4444", "#22c55e"])
        mc_date = pd.to_datetime(MC_TARGET)
        pr_date = mc_date - datetime.timedelta(days=16)

        fig.add_shape(type="line", x0=pr_date, x1=pr_date, y0=0, y1=1.10, yref="paper",
                      line=dict(color="blue", width=2, dash="dash"))
        fig.add_shape(type="line", x0=mc_date, x1=mc_date, y0=0, y1=1.01, yref="paper",
                      line=dict(color="red", width=2, dash="dash"))
        fig.add_annotation(x=pr_date, y=1.15, yref="paper", text=f"<b>Power Receiving: {pr_date.strftime('%m/%d')}</b>",
                           showarrow=False, font=dict(color="blue", size=11), xanchor="center", xshift=0)
        fig.add_annotation(x=mc_date, y=1.05, yref="paper", text=f"<b>MC: {mc_date.strftime('%m/%d')}</b>",
                           showarrow=False, font=dict(color="red", size=11), xanchor="center", xshift=0)

        for idx, row in chart_df.iterrows():
            fig.add_annotation(
                x=row["Handover"], y=row["Area"],
                text=f"<b>{row['Handover'].strftime('%m/%d')}</b>",
                showarrow=False, xanchor='right', xshift=-10,
                font=dict(size=10, color="#334155")
            )
            f_finish = pd.to_datetime(row["Pressure Test Finish"])
            f_color = "#ef4444" if f_finish > mc_date else "#334155"
            fig.add_annotation(
                x=row["Pressure Test Finish"], y=row["Area"],
                text=f"<b>{row['Pressure Test Finish'].strftime('%m/%d')}</b>",
                showarrow=False, xanchor='left', xshift=10,
                font=dict(size=10, color=f_color)
            )

        fig.update_layout(
            height=430,
            margin=dict(l=10, r=30, t=55, b=30),
            xaxis=dict(title="", tickformat="%b %Y", side="bottom"),
            yaxis=dict(title="", autorange="reversed",
                       categoryorder="array",
                       categoryarray=ORIGINAL_DISPLAY_ORDER),
            showlegend=False,
            coloraxis_showscale=False,
            plot_bgcolor="white",
            paper_bgcolor="white"
        )
        st.plotly_chart(fig, use_container_width=True)

    # --- Compact Lower Section ---
    st.markdown("""
        <style>
        .compact-section { margin-top: -70px; }
        .stDivider { margin-top: 0px; margin-bottom: 0px; }
        .flow-box { 
            background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; 
            padding: 10px; font-size: 1.05rem; text-align: center; height: 100%;
        }
        </style>
    """, unsafe_allow_html=True)

    st.markdown('<div class="compact-section">', unsafe_allow_html=True)
    # Divider removed to save space
    
    st.subheader("🔄 Construction Work Flow")
    f1, f2, f3, f4, f5, f6 = st.columns(6)
    f1.markdown('<div class="flow-box"><b>1. Fabrication</b><br>Spool & Support Fabrication</div>', unsafe_allow_html=True)
    f2.markdown('<div class="flow-box"><b>2. Structure</b><br>Handover</div>', unsafe_allow_html=True)
    f3.markdown('<div class="flow-box"><b>3. Access</b><br>Scaffolding<br>(2 Weeks)</div>', unsafe_allow_html=True)
    f4.markdown('<div class="flow-box"><b>4. Erection</b><br>Header & Branches Piping</div>', unsafe_allow_html=True)
    f5.markdown('<div class="flow-box"><b>5. Punch & Test</b><br>Pressure Test & Punch<br>(2 Weeks)</div>', unsafe_allow_html=True)
    f6.markdown('<div class="flow-box"><b>6. Pre-Commissioning</b><br>CCW Flushing & FG Pig Cleaning<br>(2 Months)</div>', unsafe_allow_html=True)

    st.subheader("✅ Essential Systems for GT #11 Start-up")

    # Matrix - Original Essential Systems (GT #11 scope only)
    df_ess = pd.DataFrame(ORIGINAL_ESSENTIAL_SYSTEMS)
    
    # Calculate Total
    total_row = pd.DataFrame([{
        "System": "TOTAL",
        "Area": "",
        "Description": "",
        "DI": int(df_ess["DI"].sum()),
        "EA": int(df_ess["EA"].sum()),
        "Criticality": "",
        "Equipment": "",
        "Remark": ""
    }])
    df_ess = pd.concat([df_ess, total_row], ignore_index=True)
    
    # Set index to 1, 2, 3... and empty string for TOTAL
    new_index_ess = [str(i+1) for i in range(len(df_ess)-1)] + [""]
    df_ess.index = new_index_ess
    
    # Rename columns for display
    df_ess = df_ess.rename(columns={"Equipment": "Relevant Equipment"})
    
    st.table(df_ess)

    st.subheader("💡 Infrastructure Dependency & Solution")
    st.markdown(f"""
        <div class="solution-card">
            <div class="solution-content">
                <div style="margin-bottom: 8pt;">1. <b>Safety Priority</b>: Prioritize Fire Fighting & HVAC in electrical rooms to allow Energizing without fire/overheating risk (by INTEGRA)</div>
                <div style="margin-bottom: 8pt;">2. <b>Power Receiving</b>: Complete AIS/Transformer & Cable installation to enable Power Receiving before CCW Flushing.</div>
                <div style="margin-bottom: 8pt;">3. <b>Structural Early Handover</b>: Accelerate Pipe Rack & MB Structure delivery through additional manpower and crash work.</div>
                <div>4. <b>Double Shift</b>: Implement 24/7 welding for Main Building Structural Header-to-Branch transitions.</div>
            </div>
        </div>
    """, unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
# --- Technical Report(1) Tab ---
with tab_rep1:
    # Header row with Title and Print button
    h_c1, h_c2 = st.columns([8.5, 1.5])
    with h_c1:
        st.markdown("<h1 style='margin-bottom:0; padding-bottom:0;'>GT #11 EARLY POWER PIPING REPORT</h1>", unsafe_allow_html=True)
        st.markdown("<hr style='margin: 0; padding: 0;'>", unsafe_allow_html=True)
    with h_c2:
        st.markdown("<div style='text-align:right; margin-top:20px;'>", unsafe_allow_html=True)
        add_print_button()
        st.markdown("</div>", unsafe_allow_html=True)
        
    st.markdown("""
    <div class="report-card">
    
    
    
    ## 1. Project Sequence & Major Constraints
    The GT #11 startup is governed not only by piping progress but also by critical infrastructure and safety prerequisites.
    
    ### A. Permanent Power Receiving
    Power Receiving is a **non-negotiable milestone** that must occur before the operation of all BOP equipment, including CCW Pumps and Air Compressors.
    - **Pre-requisites**: Installation of AIS (Air Insulated Switchgear), Transformers, Electrical Panels, and Main Cabling must be finalized.
    - **Safety Priority**: Energizing cannot proceed without the completion of **Fire Fighting (FF)** and **HVAC** systems to mitigate the risk of fire and overheating.
    - **Impact**: Power Receiving must be completed *prior* to the CCW Flushing phase to ensure pump functionality.

    ### B. Major Bottleneck Section: Main Building Steel Structure
    The Main Building Steel Structure remains the definitive piping bottleneck, as branch piping distributions depend on MB Structure certification.
    
    ## 2. Pre-Commissioning Activity: CCW & Fuel Gas System
    - **Prior to Pre-Commissioning**: Formal completion of piping, supports, punch clearance, and pressure tests.
    - **Pre-Commissioning Readiness**: Pre-fabricate temporary items (Launchers, Receivers) for immediate post-test installation.
    - **Pre-Commissioning**: The 2-month window post-MC is driven by high-velocity flushing of the CCW and clean-pigging of the FG lines.
    
    ## 3. GT Anti-Icing System (Winter Operation)
    For the target operation date of **Dec 31, 2026**, the Anti-Icing system is mandatory to prevent GT Air Intake freezing.
    - **System Logic**: Utilizes Steam from the Aux. Boiler to heat water, which is then circulated via Hot Water (HW) Supply Pumps and Heat Exchangers.
    - **Piping Routing**: The AS/HW piping path encompasses: Aux. Boiler Building & HW Pump House → PR #4 → PR #3 → MB Structure → GT #11.
    - **Work Scope**: Installation of Unit B0 and B1 piping volumes is mandatory for this system to be functional for the winter startup.
    
    ## 4. Resource Allocation Issue
    - **Workspace Constraint**: Manpower in Main Building Structure remains capped at **6 teams** due to elevation and space safety constraints.
    - **Dynamic Redistribution**: Teams will be prioritized for AS/HW and FG/CCW essential systems to meet the Dec 31 milestone.
    
    ## 5. Essential Systems & Criticality Matrix
    
    | Essential System | Area | Description | Piping (DI) | Support (EA) | Criticality | Relevant Equipment | Remark |
    | :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- |
    | **1. Fuel Gas & FGH** | MB/GT/FGH | Supply Fuel Gas to GT | 2,480 | 610 | **Mandatory** | FGSS System, Performance Heater & Final Filter | |
    | **2. CCW** | MB/PR/PH1 | Supply & Return of CCW for equipment cooling down | 4,120 | 980 | **Mandatory** | PH#1 & FFC#1 Integrated | |
    | **3. IA** | All | Supply Instrument Air to all plant | 1,150 | 290 | **Operation** | Air Compressors, IA Dryers | |
    | **4. AS/HW** | PR#3/4/AB | supply & return of Hot Water for GT Anti-Icing | 2,650 | 660 | **Mandatory** | Aux. Boiler, Hot Water Supply Pump, Heat Exchanger | |
    | **5. N2** | GT / PR | Purge & Vent for Fuel Gas | 320 | 90 | **Mandatory** | N2 Storage Tank, N2 Supply System | |
    | **6. GT MISC**| MB | GT Lube Oil Mist & Fuel Gas Vent | 650 | 180 | **Mandatory** | GT Enclosure, Vent Fans | |
    | **7. Demi. Water**| WT / PR | Supply Demi. Water to plant | 2,000 | 480 | **Highest** | DW Tank & Pump Station | **Under review for sourcing from other power plant (by INTEGRA)** |
    | **TOTAL** | - | - | **13,370** | **3,290** | - | - | - |
    
    > **⚠️ Volumetric Note**: Total Area Construction Volume (**29,534 DI**) differs from the **Essential Start-up Scope** (**13,370 DI**) documented above. 
    > 1. **Focus Area**: Reflecting only Unit B0 (Common) and B1 (Main) systems required for the Dec 31 Milestone.
    > 2. **Field Erection Only**: DI values represent site welding/installation only (excl. Shop Fab).
    > 3. **Volume Update**: Current figures reflect the official B0/B1 integration based on the latest Joint Master.
    
    - **Integrated Approach**: Welding and NDT for these systems are prioritized for the Main Building and Pipe Rack #3/#4 paths.
    
    ## 6. Action Items for On-time Completion
    
    1. **Safety Priority**: Prioritize Fire Fighting & HVAC in electrical rooms to allow Energizing without fire/overheating risk.
    2. **Power Receiving**: Complete AIS/Transformer & Cable installation to enable Power Receiving before CCW Flushing.
    3. **Structural Early Handover**: Accelerate Pipe Rack & MB Structure delivery through additional manpower and crash work.
    4. **Double Shift**: Implement 24/7 welding for Header-to-Branch transitions to maximize daily output in the Main Building.

    ## 7. Project Risk & Implementation Constraints
    
    ### A. Piping Isolation & Loop Integrity (Flushing Requirements)
    - GT #11 Early Power requires functional MB utility lines. The integrated design normally requires all units to be interconnected.
    - Without physical isolation, mediums will disperse through Units #2~4 header paths, rendering independent pressurized flushing impossible.
    - Installation of **Blind Flanges or Isolation Valves** at Unit #1 Tees is mandatory to create a temporary **Closed System** for loop integrity.

    ### B. Operational Restrictions (Hot Work Prohibition)
    - Introduction of Fuel Gas (FG) for GT #11 operation will trigger strict safe-work protocols in the Main Building.
    - Hot-Work Restrictions will indefinitely suspend the installation of structural and piping systems for Units #2~4.
    - Completion of all header welding is required prior to gas admission to define and manage 'Dead Zone' for remaining expansion work.

    ### C. Impact on Total Project Milestone (Design Sync Risk)
    - Isolating Unit #1 for early startup is a significant deviation from the core integrated Power Plant Design Concept.
    - Severe resource inefficiencies and spatial conflicts will cause a **minimum 3-month delay** in total project completion.
    - Coordination with the client regarding potential schedule extensions and technical requirements for final unit integration.



    </div>
    """, unsafe_allow_html=True)

# --- Revision Tab: GT #11 + GT #12 Combined Analysis ---
with tab_rev:
    # --- Header ---
    rv_c1, rv_c2 = st.columns([8.2, 1.8])
    with rv_c1:
        st.markdown('<h1 style="font-size: 1.8rem; margin: 0; padding: 0;">📊 Revision: GT #11 + GT #12 Combined Scope Analysis</h1>', unsafe_allow_html=True)
    with rv_c2:
        add_print_button()

    st.markdown(f"""
        <div style="background-color:#fff3e0; padding:15px; border-radius:10px; border-left:8px solid #ff9800; margin-bottom:10px;">
            <span style="font-size:1.15rem; font-weight:bold; color:#e65100;">
                🚨 Ultimate Bottleneck in timeline : {rev_math_bottleneck} (Structure Access for header piping installation) | Target Float: {rev_net_float} Days
            </span>
        </div>
        <div style="background-color:#e0f2f1; padding:15px; border-radius:10px; border-left:8px solid #00897b; margin-bottom:20px;">
            <span style="font-size:1.15rem; font-weight:bold; color:#004d40;">
                ⚠️ Major Constraint: Power Receiving must be completed before CCW/HW Pump operation.
                Energizing requires FF (Fire Fighting) &amp; HVAC installation to prevent overheating/fire.
            </span>
        </div>
    """, unsafe_allow_html=True)

    # --- KPI Grid ---
    rm1, rm2, rm3, rm4, rm5, rm6 = st.columns(6)
    rm1.metric("Piping Vol.", f"{REVISION_TOTAL_DI:,} DI")
    rm2.metric("Support Vol.", f"{REVISION_TOTAL_EA:,} EA")
    rm3.metric("Avg Field DI/DAY", f"{prod_di} DI")
    rm4.metric("Avg Support EA/DAY", f"{prod_ea:.1f} EA")
    rm5.metric("Target Float", f"{rev_net_float} D", delta_color="normal" if rev_net_float >= 0 else "inverse")
    rm6.metric("Total Teams", f"{ep_work_teams:.1f}")

    st.divider()

    # --- Table + Chart ---
    rl, rr = st.columns([4.5, 5.5])
    with rl:
        st.subheader("📊 Detailed Milestone Forecast")
        st.markdown('<div style="height: 40px;"></div>', unsafe_allow_html=True)
        rev_ordered = rev_df.set_index("Area").reindex(AREA_DISPLAY_ORDER).reset_index()
        rev_df_view = rev_ordered.copy()
        rev_df_view["Work Teams Temp"] = rev_df_view["Work Teams"].map(lambda x: f"{x:.1f}")
        for c in ["Handover", "Pressure Test Finish"]:
            rev_df_view[c] = rev_df_view[c].apply(lambda x: x.strftime('%Y-%m-%d'))

        rev_display_cols = ["Area", "Work Teams Temp", "Handover", "Piping(DI)", "Support(EA)", "Pressure Test Finish", "Float"]
        rev_df_display = rev_df_view[rev_display_cols].rename(columns={"Work Teams Temp": "Work Teams"})

        rev_total_row = pd.DataFrame([{
            "Area": "TOTAL",
            "Work Teams": f"{rev_df['Work Teams'].sum():.1f}",
            "Handover": "",
            "Piping(DI)": int(rev_df["Piping(DI)"].sum()),
            "Support(EA)": int(rev_df["Support(EA)"].sum()),
            "Pressure Test Finish": "",
            "Float": int(rev_df["Float"].min())
        }])
        rev_df_display = pd.concat([rev_df_display, rev_total_row], ignore_index=True)
        rev_new_index = [str(i + 1) for i in range(len(rev_df_display) - 1)] + [""]
        rev_df_display.index = rev_new_index
        st.table(rev_df_display)

        excel_buf = create_revision_excel(rev_df, MC_TARGET)
        st.download_button(
            label="📥 Download Excel (A3)",
            data=excel_buf,
            file_name="Revision_Milestone_Forecast.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    with rr:
        st.subheader("📅 Construction Path")
        rev_chart_df = rev_df.set_index("Area").reindex(AREA_DISPLAY_ORDER).reset_index()
        rev_fig = px.timeline(
            rev_chart_df, x_start="Handover", x_end="Pressure Test Finish", y="Area",
            color="Float", color_continuous_scale=["#ef4444", "#22c55e"]
        )
        mc_date_rev = pd.to_datetime(MC_TARGET)
        pr_date_rev = mc_date_rev - datetime.timedelta(days=16)

        rev_fig.add_shape(type="line", x0=pr_date_rev, x1=pr_date_rev, y0=0, y1=1.10, yref="paper",
                          line=dict(color="blue", width=2, dash="dash"))
        rev_fig.add_shape(type="line", x0=mc_date_rev, x1=mc_date_rev, y0=0, y1=1.01, yref="paper",
                          line=dict(color="red", width=2, dash="dash"))
        rev_fig.add_annotation(x=pr_date_rev, y=1.15, yref="paper",
                               text=f"<b>Power Receiving: {pr_date_rev.strftime('%m/%d')}</b>",
                               showarrow=False, font=dict(color="blue", size=11), xanchor="center")
        rev_fig.add_annotation(x=mc_date_rev, y=1.05, yref="paper",
                               text=f"<b>MC: {mc_date_rev.strftime('%m/%d')}</b>",
                               showarrow=False, font=dict(color="red", size=11), xanchor="center")

        for idx, row in rev_chart_df.iterrows():
            rev_fig.add_annotation(
                x=row["Handover"], y=row["Area"],
                text=f"<b>{row['Handover'].strftime('%m/%d')}</b>",
                showarrow=False, xanchor="right", xshift=-10,
                font=dict(size=10, color="#334155")
            )
            f_finish_rev = pd.to_datetime(row["Pressure Test Finish"])
            f_color_rev = "#ef4444" if f_finish_rev > mc_date_rev else "#334155"
            rev_fig.add_annotation(
                x=row["Pressure Test Finish"], y=row["Area"],
                text=f"<b>{row['Pressure Test Finish'].strftime('%m/%d')}</b>",
                showarrow=False, xanchor="left", xshift=10,
                font=dict(size=10, color=f_color_rev)
            )

        rev_fig.update_layout(
            height=500,
            margin=dict(l=10, r=30, t=55, b=30),
            xaxis=dict(title="", tickformat="%b %Y", side="bottom"),
            yaxis=dict(title="", autorange="reversed",
                       categoryorder="array",
                       categoryarray=AREA_DISPLAY_ORDER),
            showlegend=False,
            coloraxis_showscale=False,
            plot_bgcolor="white",
            paper_bgcolor="white"
        )
        st.plotly_chart(rev_fig, use_container_width=True)

    # --- Construction Work Flow ---
    st.markdown('<div class="compact-section">', unsafe_allow_html=True)

    st.subheader("🔄 Construction Work Flow")
    rf1, rf2, rf3, rf4, rf5, rf6 = st.columns(6)
    rf1.markdown('<div class="flow-box"><b>1. Fabrication</b><br>Spool & Support Fabrication</div>', unsafe_allow_html=True)
    rf2.markdown('<div class="flow-box"><b>2. Structure</b><br>Handover</div>', unsafe_allow_html=True)
    rf3.markdown('<div class="flow-box"><b>3. Access</b><br>Scaffolding<br>(2 Weeks)</div>', unsafe_allow_html=True)
    rf4.markdown('<div class="flow-box"><b>4. Erection</b><br>Header & Branches Piping</div>', unsafe_allow_html=True)
    rf5.markdown('<div class="flow-box"><b>5. Punch & Test</b><br>Pressure Test & Punch<br>(2 Weeks)</div>', unsafe_allow_html=True)
    rf6.markdown('<div class="flow-box"><b>6. Pre-Commissioning</b><br>CCW Flushing & FG Pig Cleaning<br>(2 Months)</div>', unsafe_allow_html=True)

    st.subheader("✅ Essential Systems for GT #11 & GT #12 Start-up (Revised)")

    df_ess_rev = pd.DataFrame(ESSENTIAL_SYSTEMS)

    total_row_rev = pd.DataFrame([{
        "System": "TOTAL",
        "Area": "",
        "Description": "",
        "DI": int(df_ess_rev["DI"].sum()),
        "EA": int(df_ess_rev["EA"].sum()),
        "Criticality": "",
        "Equipment": "",
        "Remark": ""
    }])
    df_ess_rev = pd.concat([df_ess_rev, total_row_rev], ignore_index=True)
    new_index_ess_rev = [str(i+1) for i in range(len(df_ess_rev)-1)] + [""]
    df_ess_rev.index = new_index_ess_rev
    df_ess_rev = df_ess_rev.rename(columns={"Equipment": "Relevant Equipment"})
    st.table(df_ess_rev)

    st.subheader("💡 Infrastructure Dependency & Solution")
    st.markdown(f"""
        <div class="solution-card">
            <div class="solution-content">
                <div style="margin-bottom: 8pt;">1. <b>Safety Priority</b>: Prioritize Fire Fighting & HVAC in electrical rooms to allow Energizing without fire/overheating risk (by INTEGRA)</div>
                <div style="margin-bottom: 8pt;">2. <b>Power Receiving</b>: Complete AIS/Transformer & Cable installation to enable Power Receiving before CCW Flushing.</div>
                <div style="margin-bottom: 8pt;">3. <b>Structural Early Handover</b>: Accelerate Pipe Rack & MB Structure delivery through additional manpower and crash work.</div>
                <div>4. <b>Double Shift</b>: Implement 24/7 welding for Main Building Structural Header-to-Branch transitions.</div>
            </div>
        </div>
    """, unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

# --- Technical Report(2) Tab ---
with tab_rep2:
    # Header row with Title and Print button
    h_c1, h_c2 = st.columns([8.5, 1.5])
    with h_c1:
        st.markdown("<h1 style='margin-bottom:0; padding-bottom:0;'>GT #11 EARLY POWER PIPING REPORT</h1>", unsafe_allow_html=True)
        st.markdown("<hr style='margin: 0; padding: 0;'>", unsafe_allow_html=True)
    with h_c2:
        st.markdown("<div style='text-align:right; margin-top:20px;'>", unsafe_allow_html=True)
        add_print_button()
        st.markdown("</div>", unsafe_allow_html=True)
        
    st.markdown("""
    <div class="report-card">
    
    
    
    ## 1. Project Sequence & Major Constraints
    The GT #11 startup is governed not only by piping progress but also by critical infrastructure and safety prerequisites.
    
    ### A. Permanent Power Receiving
    Power Receiving is a **non-negotiable milestone** that must occur before the operation of all BOP equipment, including CCW Pumps and Air Compressors.
    - **Pre-requisites**: Installation of AIS (Air Insulated Switchgear), Transformers, Electrical Panels, and Main Cabling must be finalized.
    - **Safety Priority**: Energizing cannot proceed without the completion of **Fire Fighting (FF)** and **HVAC** systems to mitigate the risk of fire and overheating.
    - **Impact**: Power Receiving must be completed *prior* to the CCW Flushing phase to ensure pump functionality.

    ### B. Major Bottleneck Section: Main Building Steel Structure
    The Main Building Steel Structure remains the definitive piping bottleneck, as branch piping distributions depend on MB Structure certification.
    
    ## 2. Pre-Commissioning Activity: CCW & Fuel Gas System
    - **Prior to Pre-Commissioning**: Formal completion of piping, supports, punch clearance, and pressure tests.
    - **Pre-Commissioning Readiness**: Pre-fabricate temporary items (Launchers, Receivers) for immediate post-test installation.
    - **Pre-Commissioning**: The 2-month window post-MC is driven by high-velocity flushing of the CCW and clean-pigging of the FG lines.
    
    ## 3. GT Anti-Icing System (Winter Operation)
    For the target operation date of **Dec 31, 2026**, the Anti-Icing system is mandatory to prevent GT Air Intake freezing.
    - **System Logic**: Utilizes Steam from the Aux. Boiler to heat water, which is then circulated via Hot Water (HW) Supply Pumps and Heat Exchangers.
    - **Piping Routing**: The AS/HW piping path encompasses: Aux. Boiler Building & HW Pump House → PR #4 → PR #3 → MB Structure → GT #11.
    - **Work Scope**: Installation of Unit B0 and B1 piping volumes is mandatory for this system to be functional for the winter startup.
    
    ## 4. Resource Allocation Issue
    - **Workspace Constraint**: Manpower in Main Building Structure remains capped at **6 teams** due to elevation and space safety constraints.
    - **Dynamic Redistribution**: Teams will be prioritized for AS/HW and FG/CCW essential systems to meet the Dec 31 milestone.
    
    ## 5. Essential Systems & Criticality Matrix
    
    | Essential System | Area | Description | Piping (DI) | Support (EA) | Criticality | Relevant Equipment | Remark |
    | :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- |
    | **1. Fuel Gas & FGH** | MB/GT/FGH | Supply Fuel Gas to GT | 2,480 | 610 | **Mandatory** | FGSS System, Performance Heater & Final Filter | |
    | **2. CCW** | MB/PR/PH1 | Supply & Return of CCW for equipment cooling down | 4,120 | 980 | **Mandatory** | PH#1 & FFC#1 Integrated | |
    | **3. IA** | All | Supply Instrument Air to all plant | 1,150 | 290 | **Operation** | Air Compressors, IA Dryers | |
    | **4. AS/HW** | PR#3/4/AB | supply & return of Hot Water for GT Anti-Icing | 2,650 | 660 | **Mandatory** | Aux. Boiler, Hot Water Supply Pump, Heat Exchanger | |
    | **5. N2** | GT / PR | Purge & Vent for Fuel Gas | 320 | 90 | **Mandatory** | N2 Storage Tank, N2 Supply System | |
    | **6. GT MISC**| MB | GT Lube Oil Mist & Fuel Gas Vent | 650 | 180 | **Mandatory** | GT Enclosure, Vent Fans | |
    | **7. Demi. Water**| WT / PR | Supply Demi. Water to plant | 2,000 | 480 | **Highest** | DW Tank & Pump Station | **Under review for sourcing from other power plant (by INTEGRA)** |
    | **TOTAL** | - | - | **13,370** | **3,290** | - | - | - |
    
    > **⚠️ Volumetric Note**: Total Area Construction Volume (**29,534 DI**) differs from the **Essential Start-up Scope** (**13,370 DI**) documented above. 
    > 1. **Focus Area**: Reflecting only Unit B0 (Common) and B1 (Main) systems required for the Dec 31 Milestone.
    > 2. **Field Erection Only**: DI values represent site welding/installation only (excl. Shop Fab).
    > 3. **Volume Update**: Current figures reflect the official B0/B1 integration based on the latest Joint Master.
    
    - **Integrated Approach**: Welding and NDT for these systems are prioritized for the Main Building and Pipe Rack #3/#4 paths.
    
    ## 6. Action Items for On-time Completion
    
    1. **Safety Priority**: Prioritize Fire Fighting & HVAC in electrical rooms to allow Energizing without fire/overheating risk.
    2. **Power Receiving**: Complete AIS/Transformer & Cable installation to enable Power Receiving before CCW Flushing.
    3. **Structural Early Handover**: Accelerate Pipe Rack & MB Structure delivery through additional manpower and crash work.
    4. **Double Shift**: Implement 24/7 welding for Header-to-Branch transitions to maximize daily output in the Main Building.

    ## 7. Project Risk & Implementation Constraints
    
    ### A. Piping Isolation & Loop Integrity (Flushing Requirements)
    - GT #11 Early Power requires functional MB utility lines. The integrated design normally requires all units to be interconnected.
    - Without physical isolation, mediums will disperse through Units #2~4 header paths, rendering independent pressurized flushing impossible.
    - Installation of **Blind Flanges or Isolation Valves** at Unit #1 Tees is mandatory to create a temporary **Closed System** for loop integrity.

    ### B. Operational Restrictions (Hot Work Prohibition)
    - Introduction of Fuel Gas (FG) for GT #11 operation will trigger strict safe-work protocols in the Main Building.
    - Hot-Work Restrictions will indefinitely suspend the installation of structural and piping systems for Units #2~4.
    - Completion of all header welding is required prior to gas admission to define and manage 'Dead Zone' for remaining expansion work.

    ### C. Impact on Total Project Milestone (Design Sync Risk)
    - Isolating Unit #1 for early startup is a significant deviation from the core integrated Power Plant Design Concept.
    - Severe resource inefficiencies and spatial conflicts will cause a **minimum 3-month delay** in total project completion.
    - Coordination with the client regarding potential schedule extensions and technical requirements for final unit integration.



    </div>
    """, unsafe_allow_html=True)

